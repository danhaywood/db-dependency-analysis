import argparse
import os

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from scipy.cluster.hierarchy import linkage, leaves_list
from scipy.spatial.distance import pdist
from sklearn.decomposition import PCA
from sklearn.manifold import TSNE
from scipy.cluster.hierarchy import dendrogram
import plotly.express as px
import networkx as nx
import community as community_louvain  # this is python-louvain
from pyvis.network import Network
import igraph as ig
import leidenalg

def load_fk_data(file_path):
    df = pd.read_excel(file_path)
    df.columns = [col.strip().lower() for col in df.columns]
    return df


def build_matrix(df):
    tables = sorted(set(df['from_table']) | set(df['to_table']))
    matrix = pd.DataFrame('', index=tables, columns=tables)

    fk_set = set((row['from_table'], row['to_table']) for _, row in df.iterrows())

    for from_table, to_table in fk_set:
        if (to_table, from_table) in fk_set:
            # Bidirectional FK
            matrix.at[from_table, to_table] = 'Y'
            matrix.at[to_table, from_table] = 'Y'
        else:
            # Unidirectional
            matrix.at[from_table, to_table] = 'Y'

    return matrix

def plot_projection(clustered_tables, coords, output_base, title, suffix, ndepend_style):
    plt.figure(figsize=(12, 10))
    x, y = coords[:, 0], coords[:, 1]

    for i, table in enumerate(clustered_tables):
        plt.scatter(x[i], y[i], s=30, color='steelblue')
        label = table.split('.')[-1][:20]  # Truncate long names
        plt.text(x[i], y[i], label, fontsize=8, alpha=0.7)

    plt.title(title, fontsize=14)
    plt.grid(True)
    plt.tight_layout()

    if ndepend_style:
        output_path = f"{output_base}.{suffix}.ndepend-style.png"
    else:
        output_path = f"{output_base}.{suffix}.png"

    plt.savefig(output_path, dpi=150)
    print(f"📊 Plot saved to: {output_path}")

def plot_dendrogram(linkage_matrix, labels, output_base, ndepend_style):
    plt.figure(figsize=(14, 8))
    dendrogram(linkage_matrix, labels=[lbl.split('.')[-1][:30] for lbl in labels], leaf_rotation=90)
    plt.title("Hierarchical Clustering Dendrogram")
    plt.tight_layout()

    if ndepend_style:
        output_path = f"{output_base}.hierarchical.ndepend-style.png"
    else:
        output_path = f"{output_base}.hierarchical.png"

    plt.savefig(output_path, dpi=150)
    print(f"🌳 Dendrogram saved to: {output_path}")

def plot_interactive_projection(clustered_tables, coords, output_base, algorithm, ndepend_style):
    df = pd.DataFrame(coords, columns=["PC1", "PC2"])
    df["Table"] = clustered_tables
    df["Schema"] = [t.split(".")[0] if "." in t else "default" for t in clustered_tables]
    df["Label"] = [t.split(".")[-1] for t in clustered_tables]

    fig = px.scatter(
        df, x="PC1", y="PC2",
        hover_name="Table",
        color="Schema",
        text="Label",
        title=f"{algorithm.upper()} Projection of Table Dependencies (Interactive)",
        width=1000,
        height=700
    )

    fig.update_traces(textposition='top center', marker=dict(size=10, opacity=0.7))
    fig.update_layout(legend_title_text='Schema')

    if ndepend_style:
        output_path = f"{output_base}.{algorithm}.ndepend-style.html"
    else:
        output_path = f"{output_base}.{algorithm}.html"

    fig.write_html(output_path)
    print(f"🌐 Interactive PCA plot saved to: {output_path}")

def plot_louvain_network(df, partition, output_base):
    import plotly.graph_objs as go

    edges = [(row['from_table'], row['to_table']) for _, row in df.iterrows()]
    G = nx.Graph()
    G.add_edges_from(edges)

    pos = nx.spring_layout(G, seed=42)
    node_trace = go.Scatter(
        x=[pos[n][0] for n in G.nodes],
        y=[pos[n][1] for n in G.nodes],
        mode='markers+text',
        text=[n.split('.')[-1] for n in G.nodes],
        hovertext=[n for n in G.nodes],
        textposition="top center",
        marker=dict(
            size=10,
            color=[partition.get(n, -1) for n in G.nodes],
            colorscale='Viridis',
            showscale=True,
            colorbar=dict(title="Community")
        )
    )

    edge_trace = []
    for src, tgt in G.edges:
        edge_trace.append(go.Scatter(
            x=[pos[src][0], pos[tgt][0], None],
            y=[pos[src][1], pos[tgt][1], None],
            mode='lines',
            line=dict(width=0.5, color='#888'),
            hoverinfo='none'
        ))

    fig = go.Figure(data=edge_trace + [node_trace])
    fig.update_layout(
        title="Louvain Clustering of FK Graph",
        showlegend=False,
        width=1000,
        height=800
    )
    output_path = f"{output_base}.louvain.html"
    fig.write_html(output_path)
    print(f"🌐 Louvain network plot saved to: {output_path}")


def plot_lxxxn_pyvis(df, partition, output_base, algorithm):
    from pyvis.network import Network
    import networkx as nx

    edges = [(row['from_table'], row['to_table']) for _, row in df.iterrows()]
    G = nx.Graph()
    G.add_edges_from(edges)

    net = Network(height='800px', width='100%', notebook=False, directed=False)
    net.from_nx(G)

    for node in net.nodes:
        name = node['id']
        degree = G.degree(name)
        community = partition.get(name, -1)

        node['title'] = f"{name}\nCommunity: {community}\nFK degree: {degree}"
        node['group'] = community

    # Optional: tweak physics settings
    net.set_options("""
    var options = {
      "nodes": {
        "shape": "dot",
        "size": 10,
        "font": {"size": 12}
      },
      "edges": {
        "color": {"inherit": true},
        "smooth": false
      },
      "physics": {
        "forceAtlas2Based": {
          "gravitationalConstant": -50,
          "springLength": 100,
          "springConstant": 0.08
        },
        "minVelocity": 0.75,
        "solver": "forceAtlas2Based"
      }
    }
    """)

    output_path = f"{output_base}.{algorithm}.anim.html"
    net.write_html(output_path)
    print(f"🎞️ Animated {algorithm} graph saved to: {output_path}")


def reorder_matrix(matrix, algorithm="none", min_fks=1, df=None, threshold=2):
    if algorithm == "none":
        order = sorted(matrix.index)
        return order, None, None

    binary_matrix = matrix.replace({'Y': 1, '': 0}).astype(int)
    fk_activity = binary_matrix.sum(axis=1) + binary_matrix.sum(axis=0)
    active_tables = fk_activity[fk_activity >= min_fks].index.tolist()

    if len(active_tables) < 2:
        print("⚠️ Not enough connected tables for clustering. Falling back to alphabetical.")
        order = sorted(matrix.index)
        return order, None, None

    sub_matrix = binary_matrix.loc[active_tables, active_tables]

    try:
        if algorithm == "cosine":
            distance_matrix = pdist(sub_matrix.values, metric='cosine')
            if not np.isfinite(distance_matrix).all():
                raise ValueError("Non-finite values detected in distance matrix.")
            linkage_matrix = linkage(distance_matrix, method='average')
            order_indices = leaves_list(linkage_matrix)
            order = sub_matrix.index[order_indices].tolist()
            return order + sorted(set(matrix.index) - set(order)), order, None

        elif algorithm == "hierarchical":
            distance_matrix = pdist(sub_matrix.values, metric='jaccard')
            linkage_matrix = linkage(distance_matrix, method='average')
            order_indices = leaves_list(linkage_matrix)
            order = sub_matrix.index[order_indices].tolist()
            return order + sorted(set(matrix.index) - set(order)), order, linkage_matrix

        elif algorithm == "pca":
            pca = PCA(n_components=2)
            coords = pca.fit_transform(sub_matrix.values)
            order_indices = np.argsort(coords[:, 0])
            order = sub_matrix.index[order_indices].tolist()
            return order + sorted(set(matrix.index) - set(order)), order, coords

        elif algorithm == "tsne":
            tsne = TSNE(n_components=2, random_state=42, init='pca', perplexity=30, n_iter=1000)
            coords = tsne.fit_transform(sub_matrix.values)
            order_indices = np.argsort(coords[:, 0])
            order = sub_matrix.index[order_indices].tolist()
            return order + sorted(set(matrix.index) - set(order)), order, coords

        elif algorithm == "louvain":
            # Step 1: Build full graph
            edges = [(row['from_table'], row['to_table']) for _, row in df.iterrows()]
            G_full = nx.Graph()
            G_full.add_edges_from(edges)

            # Step 2: Remove low-degree nodes (noise)
            G = G_full.copy()
            noise_threshold = threshold
            low_degree_nodes = [n for n, d in G.degree() if d < noise_threshold]
            G.remove_nodes_from(low_degree_nodes)

            # Step 3: Run Louvain clustering
            partition = community_louvain.best_partition(G)

            # Step 4: Order tables by cluster
            ordered_nodes = sorted(partition.items(), key=lambda x: (x[1], x[0]))
            clustered_tables = [t for t, _ in ordered_nodes]

            # Step 5: Append noise tables at end
            inactive_tables = sorted(set(matrix.index) - set(clustered_tables))
            full_order = clustered_tables + inactive_tables

            return full_order, clustered_tables, partition

        elif algorithm == "leiden":
            # Build undirected FK graph using igraph
            edges = [(row['from_table'], row['to_table']) for _, row in df.iterrows()]
            table_set = sorted(set(df['from_table']) | set(df['to_table']))
            table_to_index = {table: i for i, table in enumerate(table_set)}

            ig_edges = [(table_to_index[a], table_to_index[b]) for a, b in edges]
            g = ig.Graph(edges=ig_edges, directed=False)
            g.vs["name"] = table_set

            # Run Leiden community detection
            partition = leidenalg.find_partition(g, leidenalg.ModularityVertexPartition)

            # Convert partition to table → community dict
            table_to_comm = {}
            for comm_id, cluster in enumerate(partition):
                for idx in cluster:
                    table_to_comm[g.vs[idx]["name"]] = comm_id

            # Sort tables by community
            ordered_tables = sorted(table_to_comm.items(), key=lambda x: (x[1], x[0]))
            clustered_tables = [t for t, _ in ordered_tables]
            inactive_tables = sorted(set(matrix.index) - set(clustered_tables))
            full_order = clustered_tables + inactive_tables

            return full_order, clustered_tables, table_to_comm

        else:
            raise ValueError(f"Unsupported algorithm: {algorithm}")

    except Exception as e:
        print(f"⚠️ Clustering failed: {e}")
        print("🔁 Falling back to alphabetical.")
        order = sorted(matrix.index)
        return order, None, None

def format_excel(file_path):
    wb = load_workbook(filename=file_path)
    ws = wb.active

    for col in ws.iter_cols(min_row=1, max_row=1, min_col=2):
        for cell in col:
            cell.alignment = Alignment(textRotation=90)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.alignment = Alignment(horizontal='right')

    ws.freeze_panes = "B2"

    max_length = 0
    for cell in ws['A']:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions['A'].width = max_length + 1

    for col_cells in ws.iter_cols(min_row=1, max_row=1, min_col=2):
        col_letter = col_cells[0].column_letter
        ws.column_dimensions[col_letter].width = 2.67

    # Center-align FK cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(filename=file_path)
    print(f"🎨 Excel formatted: {file_path}")


def main(input_file, algorithm, min_fks, ndepend_style, threshold):
    print(f"🔍 Loading foreign keys from: {input_file}")
    df = load_fk_data(input_file)

    print("🧱 Building full FK matrix...")
    matrix = build_matrix(df)

    print(f"🔁 Applying algorithm: {algorithm}")
    order, clustered_tables, coords_or_partition = reorder_matrix(matrix, algorithm, min_fks, df, threshold)

    if ndepend_style:
        reordered_matrix = matrix.loc[order, order]
    else:
        reordered_matrix = matrix.loc[order].copy()

    input_base = os.path.splitext(os.path.basename(input_file))[0]

    if ndepend_style:
        output_file = f"{input_base}.{algorithm}.ndepend-style.xlsx"
    else:
        output_file = f"{input_base}.{algorithm}.xlsx"

    reordered_matrix.to_excel(output_file)
    format_excel(output_file)
    print(f"💾 Excel saved to: {output_file}")

    if algorithm == "louvain":
        # plot_louvain_network(df, coords_or_partition, input_base)
        plot_lxxxn_pyvis(df, coords_or_partition, input_base, algorithm)

    if algorithm == "leiden":
        plot_lxxxn_pyvis(df, coords_or_partition, input_base, algorithm)

    if algorithm == "hierarchical" and coords_or_partition is not None:
        plot_dendrogram(coords_or_partition, clustered_tables, input_base, ndepend_style)

    if algorithm in ["tsne", "pca"] and clustered_tables is not None and coords_or_partition is not None:
        plot_projection(
            clustered_tables,
            coords_or_partition,
            input_base,
            f"{algorithm.upper()} Projection of Table Dependencies",
            algorithm,
            ndepend_style
        )

    if algorithm in ["pca", "tsne"]:
        plot_interactive_projection(clustered_tables, coords, input_base, algorithm, ndepend_style)

    print("✅ Done!")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Foreign Key Analysis Tool")
    parser.add_argument("--input", type=str, default="foreign_keys.xlsx", help="Input Excel file with FK relationships")
    parser.add_argument("--algorithm", type=str, choices=["none", "hierarchical", "cosine", "pca", "tsne", "louvain", "leiden"], default="none", help="Reordering algorithm")
    parser.add_argument("--min-fks", type=int, default=1, help="Minimum total FK activity (in+out) to include in clustering")
    parser.add_argument("--ndepend-style", type=bool, default=False, help="Force same row/column ordering (NDepend-style)")
    parser.add_argument("--threshold", type=int, default=2, help="FK threshold to ignore as noise (Louvain only)")
    args = parser.parse_args()

    main(args.input, args.algorithm, args.min_fks, args.ndepend_style, args.threshold)